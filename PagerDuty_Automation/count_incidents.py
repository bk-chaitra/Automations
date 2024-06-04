import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

def fetch_services(api_token, params=None):
    headers = {
        "Accept": "application/vnd.pagerduty+json;version=2",
        "Authorization": f"Token token={api_token}"
    }
    url = "https://api.pagerduty.com/services"
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        return response.json()["services"]
    else:
        print("Failed to fetch services:", response.text)
        return []

def fetch_incidents(api_token, service_id, start_date, end_date):
    headers = {
        "Accept": "application/vnd.pagerduty+json;version=2",
        "Authorization": f"Token token={api_token}"
    }
    params = {
        "since": start_date.isoformat(),
        "until": end_date.isoformat(),
        "service_ids[]": service_id,
        "limit": 100,
        "offset": 0  # Initialize offset
    }
    incidents = []
    while True:
        url = "https://api.pagerduty.com/incidents"
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            data = response.json()
            incidents.extend(data["incidents"])
            if not data["more"]:
                break  # No more pages
            params["offset"] += params["limit"]  # Increment offset for next page
        else:
            print("Failed to fetch incidents:", response.text)
            break
    return incidents


def count_incidents_per_date(incidents):
    incidents_per_date = {}
    for incident in incidents:
        occurred_at = incident["created_at"]
        occurred_date = datetime.strptime(occurred_at, "%Y-%m-%dT%H:%M:%SZ").date()
        formatted_date = occurred_date.strftime("%b %d")  # Format the date as "MMM DD"
        if formatted_date not in incidents_per_date:
            incidents_per_date[formatted_date] = 0
        incidents_per_date[formatted_date] += 1
    return incidents_per_date

def export_to_excel(incidents_data, file_name):
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()
    except Exception as e:
        print("Error loading workbook:", e)
        return

    sheet = workbook.active
    
    # Write dates as column headers
    dates = sorted(set(date for incidents in incidents_data.values() for date in incidents.keys()))
    sheet.append(["Service Name"] + dates)

    # Write service names and incident counts
    for service_name, incidents_per_date in incidents_data.items():
        row_data = [service_name]
        for date in dates:
            row_data.append(incidents_per_date.get(date, 0))
        sheet.append(row_data)

    workbook.save(file_name)

if __name__ == "__main__":
    api_token = "e+i-GkguAd26VyVUvEXg" # your api token
    start_date = datetime(2024, 5, 28, 0, 0)  # Start date with time
    end_date = datetime(2024, 6, 3, 23, 59)   # End date with time
    file_name = "incidents-count-new.xlsx"
    params = {"limit": 100}
    services = fetch_services(api_token, params=params)
    
    incidents_data = {}  # Dictionary to store incidents data for each service
    for service in services:
        service_name = service["name"]
        service_id = service["id"]
        incidents = fetch_incidents(api_token, service_id, start_date, end_date)
        if incidents:
            incidents_per_date = count_incidents_per_date(incidents)
            incidents_data[service_name] = incidents_per_date
    
    export_to_excel(incidents_data, file_name)
    print(f"Incidents exported to {file_name}")
