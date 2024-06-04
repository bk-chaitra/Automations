import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import re

def fetch_services(api_token, params=None):
    headers = {
        "Accept": "application/vnd.pagerduty+json;version=2",
        "Authorization": f"Token token={api_token}"
    }
    url = "https://api.pagerduty.com/services"
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        return response.json()["services"]
    else:
        print("Failed to fetch services:", response.text)
        return []

def fetch_incidents(api_token, service_id, start_date, end_date):
    headers = {
        "Accept": "application/vnd.pagerduty+json;version=2",
        "Authorization": f"Token token={api_token}"
    }
    params = {
        "since": start_date.isoformat(),
        "until": end_date.isoformat(),
        "service_ids[]": service_id,
        "limit": 100,
        "offset": 0  # Initialize offset
    }
    incidents = []
    while True:
        url = "https://api.pagerduty.com/incidents"
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            data = response.json()
            incidents.extend(data["incidents"])
            if not data["more"]:
                break  # No more pages
            params["offset"] += params["limit"]  # Increment offset for next page
        else:
            print("Failed to fetch incidents:", response.text)
            break
    return incidents

def export_to_excel(incidents, service_name, file_name):
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()
    except Exception as e:
        print("Error loading workbook:", e)
        return

    sheet = workbook.create_sheet(title=service_name[:31])  # Truncate service name to 31 characters
    sheet.append(["Date", "Incident ID", "Description", "Status"])
    
    current_date = None  # Variable to track the current date
    for incident in incidents:
        occurred_at = incident["created_at"]
        occurred_date = datetime.strptime(occurred_at, "%Y-%m-%dT%H:%M:%SZ").date()
        formatted_date = occurred_date.strftime("%Y-%m-%d")  # Format the date as YYYY-MM-DD
        
        # # Check if the date has changed
        # if current_date != occurred_date:
        #     # Add an empty row before adding incidents for a new date
        #     if current_date is not None:
        #         sheet.append([])
        #     current_date = occurred_date
        
        sheet.append([formatted_date, incident["id"], incident["summary"], incident["status"]])

    workbook.save(file_name)


if __name__ == "__main__":
    api_token = "" # your api token
    # end_date = datetime.now()
    # start_date = end_date - timedelta(days=7)  # Default to the last 7 days
    start_date = datetime(2024, 5, 28, 0, 0)  # Start date with time
    end_date = datetime(2024, 6, 3, 23, 59)  # End date with time 
    file_name = "incidents-latest.xlsx"
    params = {"limit": 100}
    services = fetch_services(api_token, params=params)
    for service in services:
        service_name = service["name"]
        service_id = service["id"]
        incidents = fetch_incidents(api_token, service_id, start_date, end_date)
        if incidents:
            export_to_excel(incidents, service_name, file_name)
            print(f"Incidents for {service_name} exported to {file_name}")
